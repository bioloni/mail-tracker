import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from datetime import datetime, date, timedelta
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# Create the window
root = tk.Tk()
root.title("Email App")

# Create the input fields and upload buttons
inbox_label = tk.Label(root, text="Inbox emails:")
inbox_label.grid(row=0, column=0)
inbox_input = tk.Entry(root)
inbox_input.grid(row=0, column=1)
inbox_button = tk.Button(root, text="Upload", command=lambda: select_file(inbox_input) and check_input_fields())
inbox_button.grid(row=0, column=2)

sent_label = tk.Label(root, text="Sent emails:")
sent_label.grid(row=1, column=0)
sent_input = tk.Entry(root)
sent_input.grid(row=1, column=1)
sent_button = tk.Button(root, text="Upload", command=lambda: select_file(sent_input) and check_input_fields())
sent_button.grid(row=1, column=2)

#Build dictionary with only relevant subjects
subjects=["RevIQ Configuration Call", "Engagement Assistance Needed", "Mandatory training needed before RevIQ Configuration Call", "Reminder - Engagement Assistance Needed", "ZS RevIQ", "RevIQ Activation Date Change", "RevIQ Configuration Call Recap", "We are waiting for you!", "We missed you on the call!", "RevIQ Go-Live Tomorrow", "Your property is now live on RevIQ", "Transition to New RevIQ Coach"]

# Parse dates
date_format_1 = '%H:%M %p'
date_format_2 = '%a %I:%M %p'
date_format_3 = '%a %m/%d'
date_format_4 = '%m/%d/%Y'
# Get today's date
today = datetime.now()
# Get current day of the week
current_weekday = date.today().weekday()

# Function to open the file browser and save the selected file path
def select_file(input_field):
    filepath = filedialog.askopenfilename(title="Select file")
    input_field.delete(0, tk.END)
    input_field.insert(tk.END, os.path.basename(filepath))
    input_field.filepath = filepath
    check_input_fields()

def check_input_fields():
    if inbox_input.get() and sent_input.get():
        submit_button.config(state=tk.NORMAL)
    else:
        submit_button.config(state=tk.DISABLED)

def action_gen(x):
    grace_period=3
    date_diff=(today - x['Date']).days
    if (x['Status'] == "Received") and (date_diff >= grace_period): return "URGENT Client is expecting an answer"
    elif (x['Status'] == "Received") and (date_diff < grace_period): return "Client is expecting an answer"
    elif (x['Status'] == "Sent") and (date_diff >= grace_period): return "Follow up on the email you sent"
    elif (x['Status'] == "Sent") and (date_diff < grace_period): return "Give the client more time to answer"
    else: return "Error"

def custom_sort(x):
    if "URGENT" in x['Action']:
        return 3
    elif "Client is expecting an answer" in x['Action'] or "Follow up on the email you sent" in x['Action']:
        return 2
    else: 
        return 1

def check_date_format(date_string, date_format):
    try:
        datetime.strptime(date_string, date_format)
        return True
    except:
        return False

def update_steps(row):
    for subject in step_dict.keys():
        if subject in row['Subject'].values:
            row[step_dict[subject]] = 1
    return row

# Function to convert date formats
def convert_date_format(date_string):
    if check_date_format(date_string,date_format_1):
        return today.strftime('%m/%d/%Y')
    elif check_date_format(date_string,date_format_2):
        # Assume the day of the week corresponds to this week
        datetime_object = datetime.strptime(date_string, '%a %I:%M %p')
        # Get the date of the same day of the week of the current date
        same_weekday_date = date.today() - timedelta(days=current_weekday) + timedelta(days=datetime_object.weekday())
        # Combine the date and time
        datetime_object = datetime.combine(same_weekday_date, datetime_object.time())
        # Format datetime object as '%m/%d/%Y'
        formatted_date = datetime_object.strftime('%m/%d/%Y')
        return formatted_date
    elif check_date_format(date_string,date_format_3):
        date_string = date_string + '/' + str(today.year)
        return datetime.strptime(date_string, '%a %m/%d/%Y').strftime('%m/%d/%Y')
    elif check_date_format(date_string,date_format_4):
        return date_string
    else:
        print("Odd date {}".format(date_string))
        return "01/01/1900"

#Function to submit and process the input files
def submit_files():
    # Read the files
    inbox = pd.read_csv(inbox_input.filepath)
    sent = pd.read_csv(sent_input.filepath)
    # Add the "Email type" column
    inbox["Status"] = "Received"
    sent["Status"] = "Sent"

    # Rename the "Sent" and "Received" columns to "Date"
    inbox = inbox.rename(columns={"Received": "Date"})
    sent = sent.rename(columns={"Sent": "Date"})

    # Add the "To" and "From" columns
    inbox["To"] = "Me"
    sent["From"] = "Me"

    # Add the "Client" columns
    inbox["Client"] = inbox["From"]
    sent["Client"] = sent["To"]

    # Remove the "Size" and "Categories" columns
    inbox = inbox.drop(columns=["Size", "Categories"])
    sent = sent.drop(columns=["Size", "Categories"])

    # Concatenate the dataframes and sort by date
    global master
    master = pd.concat([inbox, sent], ignore_index=True)
    #Parse dates
    master['Date_raw']=master['Date']
    master['Date']=master['Date'].apply(lambda x: convert_date_format(x))
    master['Date'] = pd.to_datetime(master['Date'])
    
    #Checkpoint df
    master_raw = master

    # #Keep only the rows that have a relevant subject
    # master = master[master['Subject'].isin(list(step_dict.keys()))] #Keep only mails with relevant subjects (subjects are stored as dictionary keys)
    # #Add new columns to df one for each subjects step value
    # master.reindex(columns=list(step_dict.values()), fill_value=0)
    # # Group by "Client" and populate the new columns depending on the emails available for each client
    # master = master.groupby("Client").apply(update_steps)
    # # master = master.dropna()

    #Convert to string and filter
    master['Subject'] = master['Subject'].astype(str)
    master = master[master['Subject'].apply(lambda x: any(substring in x for substring in subjects))]

    # Group by "Client" and keep only the latest row
    master = master.sort_values(by='Date', ascending=False).groupby('Client').first().reset_index()

    # Add the status column
    grace_period = 3
    #Generate action column
    master['Action']=master.apply(lambda x: action_gen(x),axis=1)
    master = master.drop(columns=["From", "To"])

    #Export the df
    # Create a new column "Priority" based on the custom_sort function
    master['Priority'] = master.apply(lambda x: custom_sort(x), axis=1)
    # Sort the dataframe by the Priority column in descending order
    master = master.sort_values(by=["Priority", "Date"], ascending=[False, True])
    master = master.reset_index(drop=True)
    master['Date'] = master['Date'].dt.strftime(date_format_4)

    #Rearrange columns
    cols = ['Date','Date_raw','Client','Subject','Status','Priority','Action']
    master = master[cols]

    #Save raw file
    master.to_csv("master_raw.csv")
    # Create a new Excel file
    wb = Workbook()
    ws = wb.active

    # Define the styles for the headers and rows
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center')
    borders = Border(left=Side(border_style='thin'),
                    right=Side(border_style='thin'),
                    top=Side(border_style='thin'),
                    bottom=Side(border_style='thin'))

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Write the dataframe to the Excel file

    for i, row in master.iterrows():
        for j, cell in enumerate(row):
            c = ws.cell(row=i+2, column=j+1)
            c.value = cell
            c.border = borders
            if row.Priority == 3:
                c.fill = red_fill
            elif row.Priority == 2:
                c.fill = yellow_fill
            elif row.Priority == 1:
                c.fill = green_fill

    # Format the headers
    for col_num, value in enumerate(master.columns.values):
        c = ws.cell(row=1, column=col_num+1)
        c.value = value
        c.font = header_font
        c.alignment = header_alignment


    # Save the Excel file
    wb.save("data.xlsx")




    # Create the table
    display_table(master)


#Function to display the table     #Create the table sorting functionality
def display_table(dataframe):
    # Create the table
    
    messagebox.showinfo("Information", "Summary table was created in the same folder where you executed this script")
    columns = dataframe.columns
    table = ttk.Treeview(root, columns=columns, show='headings')
    for i in range(0,len(columns)):
        table.heading(i, text=columns[i])
        table.column(i, width=100)
    for i, row in dataframe.iterrows():
        table.insert('', 'end', values=list(row))
        table.grid(row=4, column=0, columnspan=3)

#Create submit button
submit_button = tk.Button(root, text="Submit", command=submit_files, state=tk.DISABLED)
submit_button.grid(row=2, column=1)

root.mainloop()
